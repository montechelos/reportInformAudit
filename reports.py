from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os
from collections import defaultdict

class ReportG:
   def generate_report_excel(self, data, generar_report, first_batch):
        """Genera un reporte en Excel con los datos proporcionados."""
        fecha_inicio, fecha_fin, base_datos = data

        wb = Workbook()
        wb.remove(wb.active)

        promesas_por_fecha = defaultdict(list)

        def procesar_promesas(promesas):
            """Procesa las promesas y las agrupa por fecha."""
            for promesa in promesas:
                FECHA_ACUERDO = promesa['FECHA_ACUERDO'].strftime('%Y-%m-%d')
                VALOR_ACUERDO = promesa['VALOR_ACUERDO']
                VALOR_CUOTA = promesa.get('VALOR_CUOTA', 0)

                promesas_por_fecha[FECHA_ACUERDO].append([
                    promesa['ID'],    
                    promesa['NOMBRE'],                     
                    FECHA_ACUERDO,                     
                    VALOR_ACUERDO,                     
                    promesa['NUMERO_PROMESA'],         
                    promesa['ESTADO_PROMESA'],         
                    VALOR_CUOTA,                       
                    promesa.get('ESTADO_CUOTA'),
                    promesa['CUOTAS']        
                ])

        procesar_promesas(first_batch)
        for lote_promesas in generar_report:
            procesar_promesas(lote_promesas)

        os.makedirs('reportes', exist_ok=True)
        fechas_ordenadas = sorted(promesas_por_fecha.keys())

        total_registros = 0
        for fecha in fechas_ordenadas:
            promesas = promesas_por_fecha[fecha]
            ws = wb.create_sheet(title=fecha)
            headers = ['ID', 'MONRE','FECHA ACUERDO', 'VALOR ACUERDO', 'NÚMERO PROMESA', 'ESTADO PROMESA', 'VALOR CUOTA', 'ESTADO CUOTA', 'CUOTAS']
            ws.append(headers)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)

            for promesa in promesas:
                ws.append(promesa)
                total_registros += 1

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_excel = f"reporte_{base_datos}_{timestamp}.xlsx"
        ruta = os.path.join('reportes', nombre_excel)
        wb.save(ruta)

        return nombre_excel, total_registros